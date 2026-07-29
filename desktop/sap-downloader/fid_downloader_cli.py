r"""
FID BOM Downloader (CLI version)
---------------------------------
Same SAP automation logic as fid_downloader_gui.py (IB53 search + expand/
export + convert to Excel), the only difference being this one has no
tkinter window of its own — it's a pure command-line tool built for the
lambom desktop app (Electron) to call: Electron handles the window and
progress display, this tool handles actually driving SAP.

Usage:
    fid_downloader_cli.exe <FID> [--out-dir OUTPUT_DIR]                      (default tool mode, Full BOM)
    fid_downloader_cli.exe --mode modules --so <SO> [--out-dir OUTPUT_DIR]   (Module BOM, ZOOBOM_CE_FMT)
    fid_downloader_cli.exe <FID> --mode modules [--out-dir OUTPUT_DIR]      (Module BOM, leave SO blank to resolve it from FID)
    fid_downloader_cli.exe <FID> --mode resolve                            (SO/PO lookup only, no download)
    fid_downloader_cli.exe --mode zbom --so <SO> [--out-dir OUTPUT_DIR]     (ZBOM configuration options, via VA03)

Behavior:
- Progress messages are printed to stdout one line at a time (for the caller to display live).
- tool/modules/zbom modes each produce exactly ONE clean xlsx in the end
  (intermediate files and raw multi-file output are cleaned up
  automatically): tool mode produces `<FID>.xlsx`; modules mode produces
  `<SO>_modules.xlsx`, split into sheets by module; zbom mode produces
  `<SO>_zbom.xlsx`, one flat sheet of Section/Node Key/Option Type/Option
  Selection rows. The last line on success is always
  `RESULT_PATH:<full path to the output xlsx>`, so the caller can extract
  the result file's location from this marker instead of guessing from
  other output.
- resolve mode does no download at all — it just prints `RESOLVED_SO:<so>`
  and `RESOLVED_PO:<po>` lines for the caller to parse and cache.
- Any failure prints `[Error] ...` and exits with a non-zero exit code.

Prerequisites (same as fid_downloader_gui.py):
1. A Windows machine with SAP auto-login (SSO) — no manual login required.
2. Install packages first: pip install pywin32 openpyxl
3. SAP GUI Scripting must be enabled.
4. Replace SAP_PORTAL_URL with your own company Portal's "Open SAP" link.

Packaging into an exe (run on Windows):
    pip install pyinstaller
    pyinstaller --onefile --console --name fid_downloader_cli fid_downloader_cli.py
   (Keep console output — don't add --noconsole, since Electron needs to read stdout.)
"""

import argparse
import os
import re
import shutil
import subprocess
import sys
import time
import zipfile

import win32com.client
import pythoncom
import openpyxl

# When run as a child process by Electron, stdout/stderr aren't attached to a
# real console, so Windows falls back to a codepage like cp1252 that can't
# encode Chinese/CJK text by default, crashing with UnicodeEncodeError the
# moment such a log line is printed. Force utf-8, substituting a placeholder
# for characters that still can't be encoded, so it never crashes on this.
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
sys.stderr.reconfigure(encoding="utf-8", errors="replace")


# The URL of the "Open SAP" link in the company Portal — opening it logs you
# into SAP automatically via SSO. This is YOUR personal navurl; whoever else
# uses this needs to swap in their own link.
SAP_PORTAL_URL = (
    "https://epp.fremont.lamrc.net/irj/portal"
    "?NavigationTarget=navurl://e9dc0731c19963b952ca3fbeceed6db3"
)
SAP_PROCESS_NAMES = ["saplogon.exe", "sapgui.exe"]


def log(msg):
    print(msg, flush=True)


# ---------- SAP automation logic (identical to fid_downloader_gui.py) ----------

def is_sap_process_running():
    """Check via tasklist whether any SAP-related process is running."""
    try:
        output = subprocess.check_output(
            ["tasklist"], text=True, errors="ignore",
            creationflags=subprocess.CREATE_NO_WINDOW,
        )
    except Exception:
        return False
    output_lower = output.lower()
    return any(name in output_lower for name in SAP_PROCESS_NAMES)


def kill_sap_processes():
    for name in SAP_PROCESS_NAMES:
        subprocess.run(
            ["taskkill", "/F", "/IM", name],
            capture_output=True,
            creationflags=subprocess.CREATE_NO_WINDOW,
        )
    time.sleep(2)


def launch_sap_logon():
    """Simulates what you'd normally do — click the company Portal link to open SAP: just open the URL and let the default browser + SSO handle it."""
    os.startfile(SAP_PORTAL_URL)


def try_connect_sap():
    """Try connecting and do a quick sanity check that it's actually usable; returns the session on success, None on failure."""
    try:
        sap_gui_auto = win32com.client.GetObject("SAPGUI")
        application = sap_gui_auto.GetScriptingEngine
        if application.Children.Count == 0:
            return None
        connection = application.Children(0)
        if connection.Children.Count == 0:
            return None
        session = connection.Children(0)
        _ = session.Info.SystemName  # quick check that the session actually responds
        return session
    except Exception:
        return None


def ensure_sap_connected(max_wait_seconds=120):
    """
    Ensure SAP is connected and return a usable session:
      1. First test whether there's already a usable connection; use it directly if so
      2. Can't connect but an SAP process is running -> kill it and relaunch
      3. SAP isn't open at all -> launch it directly
      4. After launching, poll and wait for auto-login (SSO) + Scripting registration to complete
    """
    log("Testing current SAP connection...")
    session = try_connect_sap()
    if session is not None:
        log("Usable connection already exists, using it directly.")
        return session

    if is_sap_process_running():
        log("Detected an SAP process running but couldn't connect, killing and relaunching...")
        kill_sap_processes()
    else:
        log("SAP isn't currently open, preparing to launch...")

    log("Launching SAP Logon...")
    launch_sap_logon()

    log("Waiting for SAP auto-login and Scripting to become ready...")
    waited = 0
    interval = 2
    while waited < max_wait_seconds:
        time.sleep(interval)
        waited += interval
        session = try_connect_sap()
        if session is not None:
            log(f"SAP connected successfully (waited about {waited}s).")
            return session
        if waited % 10 == 0:
            log(f"Still waiting for login to complete... ({waited}s elapsed)")

    raise RuntimeError(f"Waited {max_wait_seconds}s and still couldn't connect to SAP — check whether auto-login completed properly.")


def open_ib53_with_fid(session, fid):
    log("IB53: opening transaction...")
    session.findById("wnd[0]/tbar[0]/okcd").text = "/nIB53"
    session.findById("wnd[0]").sendVKey(0)
    time.sleep(1)

    session.findById("wnd[0]/usr/subENTRANCE:SAPLIBOF_R3:0100/ctxtRIBOF-IBASE").text = "0"
    session.findById("wnd[0]/usr/subENTRANCE:SAPLIBOF_R3:0100/ctxtRIBOF-IBASE").caretPosition = 1
    session.findById("wnd[0]").sendVKey(0)
    time.sleep(1)

    log(f"IB53: searching for FID {fid}...")
    session.findById("wnd[0]/usr/subENTRANCE:SAPLIBOF_R3:0100/ctxtRIBOFO-EQUNO").setFocus()
    session.findById("wnd[0]/usr/subENTRANCE:SAPLIBOF_R3:0100/ctxtRIBOFO-EQUNO").caretPosition = 0
    session.findById("wnd[0]").sendVKey(4)  # F4 -> search help
    time.sleep(1)

    # Tab 5 = "F: Equipment by technical ID number"
    session.findById("wnd[1]/usr/tabsG_SELONETABSTRIP/tabpTAB005").select()
    session.findById(
        "wnd[1]/usr/tabsG_SELONETABSTRIP/tabpTAB005"
        "/ssubSUBSCR_PRESEL:SAPLSDH4:0220/sub:SAPLSDH4:0220/txtG_SELFLD_TAB-LOW[0,24]"
    ).text = fid

    session.findById("wnd[1]/tbar[0]/btn[0]").press()
    session.findById("wnd[1]/tbar[0]/btn[0]").press()
    session.findById("wnd[0]").sendVKey(0)
    time.sleep(3)
    log("IB53: equipment loaded.")


def export_installed_base(session, save_path, filename):
    log("IB53: expanding the installed-base tree...")
    tree = session.findById("wnd[0]/shellcont/shell/shellcont[1]/shell[0]")
    tree.pressButton("IBOCX_AUFR")
    tree.pressContextButton("&PRINT_BACK")
    tree.selectContextMenuItem("&PRINT_PREV_ALL")

    log("IB53: opening the export menu...")
    session.findById("wnd[0]/mbar/menu[3]/menu[5]/menu[2]/menu[1]").select()

    session.findById(
        "wnd[1]/usr/subSUBSCREEN_STEPLOOP:SAPLSPO5:0150/sub:SAPLSPO5:0150/radSPOPLI-SELFLAG[1,0]"
    ).select()
    session.findById("wnd[1]/tbar[0]/btn[0]").press()

    log(f"IB53: saving to {save_path}\\{filename}...")
    session.findById("wnd[1]/usr/ctxtDY_PATH").text = save_path
    session.findById("wnd[1]/usr/ctxtDY_FILENAME").text = filename
    session.findById("wnd[1]/tbar[0]/btn[0]").press()

    # After saving, press F3 3 times to go back to the base screen so the
    # session returns to a clean state (matching the verified Script
    # Recording), avoiding leftover screens interfering with the Modules run
    # that follows.
    session.findById("wnd[0]").sendVKey(3)
    session.findById("wnd[0]").sendVKey(3)
    session.findById("wnd[0]").sendVKey(3)


def resolve_so_po_from_fid(session, fid):
    """
    Resolve the corresponding SO and PO from a FID — open VA03, press F4 on
    the Sales Document field, switch to the "A: Sales document according to
    customer PO number" tab, search using the FID as a wildcard (<FID>*),
    then scan the results list for the first fully-populated hit.

    This implementation is ported directly from a real, working reference
    tool ("BOM Manager", an existing internal SAP-automation app) — its
    compiled Python bytecode was disassembled to recover the exact
    algorithm, rather than guessed:
      - Pagination uses the popup's real scrollbar control
        (`wnd[1]/usr` -> `.verticalScrollbar`, `.maximum`/`.position`),
        not a guessed row-spacing constant.
      - Each candidate index (`row_idx` from 3 to 59, up to 10 pages) is
        checked at TWO fixed coordinates: `lbl[130, row_idx]` (SO-bearing
        column) and `lbl[1, row_idx]` (PO-bearing column). The FIRST
        row_idx where both are non-empty is the match (not the bottom-most
        — an earlier version of this function picked the bottom-most row,
        but the reference tool picks the first hit, and testing showed the
        reference tool's approach succeeds on FIDs (e.g. 255720) where the
        bottom-most-picking version failed with an empty-list error).
      - If no row has both fields populated, the first row with just an
        SO value is used as a fallback (PO left blank).

    Every mode that calls this re-navigates to the SO fresh afterward
    (ZOOBOM_CE_FMT / VA03 configuration both type the SO directly into
    their own field), so this never needs to leave the matched row selected
    in the popup.
    """
    session.findById("wnd[0]/tbar[0]/okcd").text = "/nVA03"
    session.findById("wnd[0]").sendVKey(0)
    time.sleep(1)
    session.findById("wnd[0]").sendVKey(4)
    time.sleep(1)

    session.findById("wnd[1]/usr/tabsG_SELONETABSTRIP/tabpTAB001").select()

    search_field_id = (
        "wnd[1]/usr/tabsG_SELONETABSTRIP/tabpTAB001"
        "/ssubSUBSCR_PRESEL:SAPLSDH4:0220/sub:SAPLSDH4:0220/txtG_SELFLD_TAB-LOW[2,24]"
    )
    fid_search = f"{fid}*"
    session.findById(search_field_id).text = fid_search
    session.findById(search_field_id).setFocus()
    session.findById(search_field_id).caretPosition = len(fid_search)
    session.findById("wnd[1]/tbar[0]/btn[0]").press()
    time.sleep(1)

    def cell_text(row, col):
        try:
            label = session.findById(f"wnd[1]/usr/lbl[{row},{col}]")
        except Exception:
            return ""
        return (label.text or "").strip()

    target_row = None
    target_so = ""
    target_po = ""
    fallback_row = None
    fallback_so = ""

    try:
        scrollbar = session.findById("wnd[1]/usr").verticalScrollbar
    except Exception:
        scrollbar = None
    scroll_max = scrollbar.maximum if scrollbar is not None else 0
    scroll_pos = 0

    for _page in range(10):
        for row_idx in range(3, 60):
            so_val = cell_text(130, row_idx)
            if not so_val:
                continue
            if fallback_row is None:
                fallback_row = row_idx
                fallback_so = so_val
            po_val = cell_text(1, row_idx)
            if not po_val:
                continue
            target_row, target_so, target_po = row_idx, so_val, po_val
            break
        if target_row is not None:
            break
        if scrollbar is None or scroll_pos >= scroll_max:
            break
        scroll_pos = min(scroll_pos + 10, scroll_max)
        scrollbar.position = scroll_pos
        time.sleep(0.5)

    if target_row is None:
        if fallback_row is None:
            raise RuntimeError("Failed to resolve SO from FID — the search results list is empty, check the SAP screen.")
        target_row, target_so = fallback_row, fallback_so

    log(f"Resolved SO={target_so!r} PO={target_po!r} from FID {fid} (row {target_row}).")

    # The F4 search-help popup (wnd[1]) is still open at this point. Whatever
    # runs next in this same SAP session (a separate CLI process attaching to
    # the same live session) starts by typing straight into wnd[0], which
    # fails with a generic "control could not be found by id" error if a
    # modal popup is still sitting on top of it — so always close it here
    # rather than leaving that to the next caller.
    try:
        session.findById("wnd[1]").close()
    except Exception:
        pass

    return target_so, target_po


def download_module_bom(session, so, out_dir):
    """
    Module BOM download — the full flow has been recorded and verified
    line-by-line using SAP GUI's Script Recording and Playback:
      1. Type /nZOOBOM_CE_FMT as the transaction code, press Enter
      2. Enter the SO into the Sales Document field (S_VBELN-LOW)
      3. Press F4 on the Folder field (P_FOLDER), a folder-selection window
         pops up; specify the output folder using the same DY_PATH field as
         the Tool BOM export, then confirm
      4. Press Execute (tbar[1]/btn[8], F8) — ZOOBOM_CE_FMT automatically
         writes multiple Excel files directly into the folder just
         specified, with no further save dialog appearing
      5. Press F3 twice to go back

    Unlike Tool BOM, we don't know in advance how many files SAP will
    produce or their exact names (ZOOBOM_CE_FMT decides that itself), so
    newly-created files are detected via the "difference in folder contents
    before vs. after execution" rather than waiting for a single fixed
    filename.
    """
    os.makedirs(out_dir, exist_ok=True)
    before = set(os.listdir(out_dir))

    session.findById("wnd[0]/tbar[0]/okcd").text = "/nZOOBOM_CE_FMT"
    session.findById("wnd[0]").sendVKey(0)
    time.sleep(1)

    session.findById("wnd[0]/usr/ctxtS_VBELN-LOW").text = so

    session.findById("wnd[0]/usr/ctxtP_FOLDER").setFocus()
    session.findById("wnd[0]/usr/ctxtP_FOLDER").caretPosition = 0
    session.findById("wnd[0]").sendVKey(4)  # F4 -> folder selection
    time.sleep(1)

    session.findById("wnd[1]/usr/ctxtDY_PATH").text = out_dir
    session.findById("wnd[1]/usr/ctxtDY_PATH").setFocus()
    session.findById("wnd[1]/usr/ctxtDY_PATH").caretPosition = len(out_dir)
    session.findById("wnd[1]/tbar[0]/btn[0]").press()
    time.sleep(1)

    session.findById("wnd[0]/tbar[1]/btn[8]").press()  # Execute (F8)
    time.sleep(3)

    session.findById("wnd[0]").sendVKey(3)
    session.findById("wnd[0]").sendVKey(3)

    log("Waiting for SAP to write the files...")
    new_files = []
    for _ in range(30):
        time.sleep(1)
        after = set(os.listdir(out_dir))
        new_files = sorted(after - before)
        if new_files:
            time.sleep(2)  # wait a bit longer to make sure files aren't still being written
            after = set(os.listdir(out_dir))
            new_files = sorted(after - before)
            break

    if not new_files:
        raise RuntimeError("Timed out waiting for new files — check whether the SAP screen is stuck, or whether the output folder is correct.")

    return [os.path.join(out_dir, name) for name in new_files]


# ---------- Module BOM raw file cleanup: coordinate fix + merge into one multi-sheet xlsx ----------
#
# Every module file produced directly by ZOOBOM_CE_FMT has broken cell
# coordinates (<c r="...">, e.g. a malformed r=" 11"-style value), which
# standard xlsx libraries (including openpyxl, used here) can't read
# directly — they raise an exception. Excel/Numbers desktop apps can open
# them fine because they're lenient, reading cells in "document order"
# instead of trusting the r= attribute — the functions below manually parse
# the xlsx's internal XML following that same logic (this is the same
# approach as lib/bom-parse.ts's readXlsxRowsLenient, just ported to
# Python); after reading out the clean data, openpyxl rewrites a fresh xlsx
# with correct coordinates, merging multiple module sheets together.


def decode_xml_entities(text):
    text = text.replace("&lt;", "<").replace("&gt;", ">").replace("&quot;", '"').replace("&apos;", "'")
    text = re.sub(r"&#x([0-9a-fA-F]+);", lambda m: chr(int(m.group(1), 16)), text)
    text = re.sub(r"&#(\d+);", lambda m: chr(int(m.group(1))), text)
    return text.replace("&amp;", "&")


def parse_shared_strings(xml_text):
    strings = []
    for m in re.finditer(r"<si>(.*?)</si>", xml_text, re.S):
        texts = re.findall(r"<t[^>]*>(.*?)</t>", m.group(1), re.S)
        strings.append(decode_xml_entities("".join(texts)))
    return strings


def parse_sheet_rows_lenient(xml_text, shared_strings):
    rows = []
    for row_match in re.finditer(r"<row\b[^>]*>(.*?)</row>", xml_text, re.S):
        cells = []
        for cell_match in re.finditer(r"<c\b([^>]*?)(?:/>|>(.*?)</c>)", row_match.group(1), re.S):
            attrs, inner = cell_match.group(1), cell_match.group(2) or ""
            type_match = re.search(r'\st="([^"]*)"', attrs)
            cell_type = type_match.group(1) if type_match else "n"
            if cell_type == "inlineStr":
                is_body = re.search(r"<is>(.*?)</is>", inner, re.S)
                texts = re.findall(r"<t[^>]*>(.*?)</t>", is_body.group(1), re.S) if is_body else []
                value = decode_xml_entities("".join(texts))
            else:
                v_match = re.search(r"<v>(.*?)</v>", inner, re.S)
                raw = v_match.group(1) if v_match else ""
                if cell_type == "s" and raw:
                    try:
                        idx = int(raw)
                        value = shared_strings[idx] if 0 <= idx < len(shared_strings) else ""
                    except ValueError:
                        value = ""
                else:
                    value = decode_xml_entities(raw) if raw else ""
            cells.append(value)
        rows.append(cells)
    return rows


def resolve_first_sheet_path(names, workbook_xml, rels_xml):
    fallback = "xl/worksheets/sheet1.xml"
    m = re.search(r'<sheet\b[^>]*\br:id="([^"]+)"', workbook_xml)
    if not m:
        return fallback
    m2 = re.search(
        rf'<Relationship\b[^>]*\bId="{re.escape(m.group(1))}"[^>]*\bTarget="([^"]+)"', rels_xml
    )
    if not m2:
        return fallback
    return f"xl/{re.sub(r'^/?xl/', '', m2.group(1))}"


def read_xlsx_rows_lenient(path):
    with zipfile.ZipFile(path) as z:
        names = z.namelist()
        workbook_xml = z.read("xl/workbook.xml").decode("utf-8", "replace") if "xl/workbook.xml" in names else ""
        rels_xml = (
            z.read("xl/_rels/workbook.xml.rels").decode("utf-8", "replace")
            if "xl/_rels/workbook.xml.rels" in names
            else ""
        )
        sheet_path = resolve_first_sheet_path(names, workbook_xml, rels_xml)
        if sheet_path not in names:
            sheet_path = "xl/worksheets/sheet1.xml"
        sheet_xml = z.read(sheet_path).decode("utf-8", "replace")
        shared_strings = (
            parse_shared_strings(z.read("xl/sharedStrings.xml").decode("utf-8", "replace"))
            if "xl/sharedStrings.xml" in names
            else []
        )
    return parse_sheet_rows_lenient(sheet_xml, shared_strings)


def collapse_spaces(value):
    return re.sub(r"\s{2,}", " ", value or "").strip()


def parse_sap_field_rows(rows):
    """Converts raw SAP field-name (MATERIAL/DESCRIPTION/UOM/BOMLEVEL/
    GENE00...) row data into a clean Part Number/Description/Level/Path/
    Qty/Unit structure, following the same rules as lib/bom-parse.ts's
    parseSapFieldRows. Returns None if the format doesn't match, in which
    case the caller keeps the raw row data instead of discarding this
    module's content entirely."""
    if len(rows) < 2:
        return None

    header = [h.strip() for h in rows[0]]

    def col_index(name):
        return header.index(name) if name in header else -1

    material_idx = col_index("MATERIAL")
    description_idx = col_index("DESCRIPTION")
    uom_idx = col_index("UOM")
    bom_qty_idx = col_index("BOM_QTY")
    newbuild_qty_idx = col_index("NEWBUILD_QTY")
    level_idx = col_index("BOMLEVEL")

    genealogy_indices = sorted(
        (int(m.group(1)), idx)
        for idx, name in enumerate(header)
        for m in [re.match(r"^GENE(\d{2,})$", name)]
        if m
    )
    genealogy_indices = [idx for _, idx in genealogy_indices]

    if material_idx == -1 or level_idx == -1 or not genealogy_indices:
        return None

    items = []
    for r in range(1, len(rows)):
        fields = rows[r]
        if not fields or all(not c.strip() for c in fields):
            continue

        def get(idx, fields=fields):
            return fields[idx].strip() if 0 <= idx < len(fields) else ""

        part_no = get(material_idx)
        level_raw = get(level_idx)
        if not part_no or not level_raw.lstrip("-").isdigit():
            continue
        level = int(level_raw)

        description = collapse_spaces(get(description_idx))
        uom = get(uom_idx) or None
        qty_raw = get(bom_qty_idx) or get(newbuild_qty_idx)
        try:
            qty = float(qty_raw) if qty_raw else None
        except ValueError:
            qty = None

        genealogy = [get(idx) for idx in genealogy_indices]
        ancestors = [g for g in genealogy[:level] if g]

        items.append(
            {
                "part_no": part_no,
                "description": description or None,
                "level": level,
                "path": "/".join(ancestors),
                "qty": qty,
                "unit": uom,
            }
        )

    return items


def derive_sheet_name(path):
    """Extracts the module's short name (GB1) from the filename (e.g.
    "R0670_130 - GB1 --- LVL02 2026-...xlsx") to use as the sheet name.
    Falls back to the full filename (without extension) if it can't be
    extracted."""
    name_no_ext = os.path.splitext(os.path.basename(path))[0]
    left = name_no_ext.split(" --- ")[0]
    parts = left.split(" - ")
    return (parts[-1].strip() if len(parts) > 1 else name_no_ext)[:31]


def merge_module_files(file_paths, out_path):
    """Merges several raw module files (whose coordinates may be broken)
    into one clean multi-sheet xlsx, one sheet per module, with columns
    normalized to Part Number/Description/Level/Path/Qty/Unit."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for path in file_paths:
        rows = read_xlsx_rows_lenient(path)
        items = parse_sap_field_rows(rows)
        ws = wb.create_sheet(title=derive_sheet_name(path))
        ws.append(["Part Number", "Description", "Level", "Path", "Qty", "Unit"])
        if items is not None:
            for item in items:
                ws.append(
                    [item["part_no"], item["description"], item["level"], item["path"], item["qty"], item["unit"]]
                )
        else:
            # Not the expected SAP field format — keep the raw row data rather than discarding this module entirely.
            for row in rows:
                ws.append(row)

    wb.save(out_path)


def parse_tab_export(txt_path):
    with open(txt_path, encoding="utf-8-sig", errors="replace") as f:
        lines = f.readlines()
    rows = []
    for line in lines:
        line = line.rstrip("\n")
        if not line.strip():
            continue
        tokens = [t.strip() for t in line.split("\t") if t.strip() != ""]
        if len(tokens) != 4:
            continue
        desc, record, qty, unit = tokens
        indent = len(line) - len(line.lstrip("\t"))
        rows.append((record, desc, indent, qty, unit))
    return rows


def write_bom_xlsx(rows, out_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BOM"
    ws.append(["Part Number", "Description", "Indent(raw tab depth)", "Qty", "Unit"])
    for record, desc, indent, qty, unit in rows:
        try:
            qty_val = float(qty)
        except ValueError:
            qty_val = qty
        ws.append([record, desc, indent, qty_val, unit])
    wb.save(out_path)


# ---------- ZBOM (SAP Variant Configuration options) ----------
#
# ZBOM is not a parts list — it's the machine's SAP LO-VC "Variant
# Configuration" characteristic/value data (Sold-To Customer, Customer
# Region, Platform Type, Position 1/2/3, etc.), reached via VA03 -> Item
# Overview -> the Configuration button, one section per configurable node
# in that order's structure. This whole section is ported from the same
# reference tool as resolve_so_po_from_fid above, using its real, verified
# control IDs and object-model calls (GuiTree/GuiTableControl), not guessed
# coordinates.

ZBOM_TABLE_ID = (
    "wnd[0]/usr/subCE_INSTANCE:SAPLCEI0:0105"
    "/subCHARACTERISTICS:SAPLCEI0:1400/tblSAPLCEI0CHARACTER_VALUES"
)
ZBOM_HEADER_ID = "wnd[0]/usr/subCE_INSTANCE:SAPLCEI0:0105/subHEADER:SAPMV45A:0460"
ZBOM_SECTION_LABEL_ID = "wnd[0]/usr/subCE_INSTANCE:SAPLCEI0:0105/subHEADER:SAPLCUKO:7035/txtRCUKO-IMAKTX"
ZBOM_CONFIG_BUTTON_ID = (
    "wnd[0]/usr/tabsTAXI_TABSTRIP_OVERVIEW/tabpT\\02/ssubSUBSCREEN_BODY:SAPMV45A:4401"
    "/subSUBSCREEN_TC:SAPMV45A:4900/subSUBSCREEN_BUTTONS:SAPMV45A:4050/btnBT_POCO"
)


def open_zbom_config(session, so):
    """Navigate to VA03 for the given SO, open Item Overview, then the
    Configuration screen for that item — leaves the session with the
    characteristic-values table (ZBOM_TABLE_ID) visible and focused.

    Each step is logged individually (rather than only logging on success at
    the end) so that if a findById call fails, the log shows exactly which
    step it happened on instead of just a generic COM error with no context —
    this whole flow is newly ported and hasn't been exercised against real
    SAP yet, so precise failure location matters more than usual here.

    Deliberately does NOT press Enter after typing the SO into
    ctxtVBAK-VBELN — a real, working SAP GUI Scripting recording of this
    exact flow (VA03 -> Item Overview -> Configuration) goes straight from
    typing the SO to pressing the Item Overview button with no Enter in
    between. An earlier version of this function did press Enter there,
    which advances SAP into the order display screen before Item Overview is
    clicked, landing on a different screen state than what
    ZBOM_CONFIG_BUTTON_ID's hardcoded tab path expects — that mismatch is
    what caused "control could not be found by id" on the Configuration
    button in real testing (FID 255720 / SO N8364)."""
    log(f"ZBOM: opening VA03 for SO {so}...")
    session.findById("wnd[0]/tbar[0]/okcd").text = "/nVA03"
    session.findById("wnd[0]").sendVKey(0)
    time.sleep(1)

    session.findById("wnd[0]").maximize()
    session.findById("wnd[0]/usr/ctxtVBAK-VBELN").text = so

    log("ZBOM: opening Item Overview...")
    session.findById("wnd[0]/tbar[1]/btn[6]").press()  # Item Overview
    time.sleep(1)

    log("ZBOM: opening Configuration...")
    session.findById(ZBOM_CONFIG_BUTTON_ID).press()  # Configuration
    time.sleep(1)

    log("ZBOM: focusing the characteristics table...")
    session.findById(f"{ZBOM_TABLE_ID}/ctxtRCTMS-MWERT[1,2]").setFocus()
    session.findById(f"{ZBOM_TABLE_ID}/ctxtRCTMS-MWERT[1,2]").caretPosition = 6


def read_zbom_table(session):
    """Reads every row of the characteristic-values table for whichever
    configuration node is currently open, scrolling via the table's real
    VerticalScrollbar (pumping COM messages so SAP actually redraws between
    scroll steps). Each cell is looked up trying the ctxt/txt/lbl prefixes
    in turn, since different characteristic types render as different
    control types. Rows where the name starts with "_____" are section
    separators, not real options."""
    table = session.findById(ZBOM_TABLE_ID)
    visible_rows = table.VisibleRowCount
    row_count = table.RowCount
    cols = [table.Columns(c).Name for c in range(table.Columns.Count)]

    seen_keys = set()
    ordered_rows = []
    scroll_pos = 0

    while scroll_pos < row_count:
        if scroll_pos > 0:
            pythoncom.PumpWaitingMessages()
            for attempt in range(3):
                if attempt > 0:
                    time.sleep(1.0)
                    pythoncom.PumpWaitingMessages()
                    table = session.findById(ZBOM_TABLE_ID)
                table.VerticalScrollbar.Position = scroll_pos
                time.sleep(0.3)
                pythoncom.PumpWaitingMessages()
                break

        for row_idx in range(visible_rows):
            if scroll_pos + row_idx >= row_count:
                break

            row_data = {}
            all_empty = True
            for c, col_name in enumerate(cols):
                cell_text = ""
                for prefix in ("ctxt", "txt", "lbl"):
                    try:
                        cell = session.findById(f"{ZBOM_TABLE_ID}/{prefix}{col_name}[{c},{row_idx}]")
                    except Exception:
                        continue
                    cell_text = cell.text
                    break
                row_data[col_name] = cell_text
                if cell_text.strip():
                    all_empty = False

            if all_empty:
                continue

            name = row_data.get("RCTMS-MNAME", "").strip()
            value = row_data.get("RCTMS-MWERT", "").strip()
            if not name or not value or name.startswith("_____"):
                continue

            row_key = (name, value)
            if row_key in seen_keys:
                continue
            seen_keys.add(row_key)
            ordered_rows.append({"option_type": name, "option_selection": value})

        scroll_pos += visible_rows

    try:
        table.VerticalScrollbar.Position = 0
    except Exception:
        pass

    return ordered_rows


def get_zbom_section_name(session, fallback):
    try:
        label = (session.findById(ZBOM_SECTION_LABEL_ID).text or "").strip()
        if label:
            return label
    except Exception:
        pass
    try:
        arktx = (session.findById(f"{ZBOM_HEADER_ID}/txtVBAP-ARKTX").text or "").strip()
        if arktx:
            return arktx
    except Exception:
        pass
    return fallback


def extract_zbom_sections(session):
    """Walks every node in the configuration tree, reading that node's
    characteristic-values table into one "section" each. Duplicate
    configurations (same first 3 options) are skipped."""
    try:
        tree = session.findById("wnd[0]/shellcont/shell")
    except Exception:
        raise RuntimeError("Tree control not found — the Configuration page doesn't appear to be open.")

    all_node_keys = list(tree.GetAllNodeKeys())
    sections = []
    seen_signatures = set()

    for node_key in all_node_keys:
        node_text = tree.GetNodeTextByKey(node_key)
        tree.selectedNode = node_key
        tree.doubleClickNode(node_key)
        time.sleep(0.3)

        rows = read_zbom_table(session)
        if not rows:
            continue

        signature = tuple((r["option_type"], r["option_selection"]) for r in rows[:3])
        if signature in seen_signatures:
            continue
        seen_signatures.add(signature)

        section_name = get_zbom_section_name(session, node_text)
        sections.append({"section": section_name, "node_key": node_key, "options": rows})

    return sections


def write_zbom_xlsx(sections, out_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ZBOM"
    ws.append(["Section", "Node Key", "Option Type", "Option Selection"])
    for entry in sections:
        for opt in entry["options"]:
            ws.append([entry["section"], entry["node_key"], opt["option_type"], opt["option_selection"]])
    wb.save(out_path)


# ---------- CLI entry point ----------

def run(fid, out_dir, mode="tool", so=None):
    if mode == "resolve":
        if not fid:
            raise RuntimeError("resolve mode requires a FID.")
        session = ensure_sap_connected()
        resolved_so, resolved_po = resolve_so_po_from_fid(session, fid)
        log(f"RESOLVED_SO:{resolved_so}")
        log(f"RESOLVED_PO:{resolved_po}")
        return None

    if mode == "zbom":
        if not so and not fid:
            raise RuntimeError("ZBOM download requires an SO or FID (if SO is omitted, it's auto-resolved from FID).")

        session = ensure_sap_connected()

        if not so:
            log(f"No SO given, resolving the corresponding SO from FID {fid}...")
            so, _po = resolve_so_po_from_fid(session, fid)

        log(f"Opening the Configuration screen for SO {so}...")
        open_zbom_config(session, so)

        log("Reading configuration options across all sections...")
        sections = extract_zbom_sections(session)
        if not sections:
            raise RuntimeError("No configuration sections found — check whether the Configuration page opened correctly.")

        total_options = sum(len(s["options"]) for s in sections)
        log(f"Found {len(sections)} section(s), {total_options} option(s) total. Writing xlsx...")
        zbom_path = os.path.join(out_dir, f"{so}_zbom.xlsx")
        os.makedirs(out_dir, exist_ok=True)
        write_zbom_xlsx(sections, zbom_path)

        log(f"Done! Wrote {zbom_path}.")
        return zbom_path

    if mode == "modules":
        if not so and not fid:
            raise RuntimeError("Module BOM download requires an SO or FID (if SO is omitted, it's auto-resolved from FID).")

        session = ensure_sap_connected()

        if not so:
            log(f"No SO given, resolving the corresponding SO from FID {fid}...")
            so, _po = resolve_so_po_from_fid(session, fid)

        # module_dir is just a scratch folder for SAP to dump raw files into;
        # once merged and cleaned up it's deleted entirely — the caller gets
        # the final single clean xlsx.
        module_dir = os.path.join(out_dir, f"{so}_modules")
        log(f"Opening ZOOBOM_CE_FMT, running with SO {so}...")
        raw_files = download_module_bom(session, so, module_dir)

        log(f"Download complete, {len(raw_files)} raw file(s), merging into one clean xlsx...")
        merged_path = os.path.join(out_dir, f"{so}_modules.xlsx")
        merge_module_files(raw_files, merged_path)
        shutil.rmtree(module_dir, ignore_errors=True)

        log(f"Done! Merged into {merged_path} (raw files cleaned up).")
        return merged_path

    os.makedirs(out_dir, exist_ok=True)
    txt_name = f"{fid}.txt"
    xlsx_name = f"{fid}.xlsx"
    txt_path = os.path.join(out_dir, txt_name)

    session = ensure_sap_connected()

    log(f"Opening IB53, searching for the equipment with FID {fid}...")
    open_ib53_with_fid(session, fid)

    log("Expanding the structure and exporting...")
    export_installed_base(session, out_dir, txt_name)

    log("Waiting for SAP to write the file...")
    for _ in range(60):
        if os.path.exists(txt_path):
            break
        time.sleep(1)
    else:
        raise RuntimeError("Timed out waiting for the exported file — check whether the SAP screen is stuck.")

    log("Parsing and converting to Excel...")
    rows = parse_tab_export(txt_path)
    if not rows:
        raise RuntimeError("Parsed 0 rows of data — open the txt file and check its format.")

    xlsx_path = os.path.join(out_dir, xlsx_name)
    write_bom_xlsx(rows, xlsx_path)

    # The txt file is just an intermediate conversion artifact — once
    # converted it's no longer needed, so remove it and keep only the final xlsx.
    os.remove(txt_path)

    log(f"Done! {len(rows)} part number(s) total.")
    return xlsx_path


def main():
    parser = argparse.ArgumentParser(description="Download a BOM from SAP by FID/SO and convert it to xlsx")
    parser.add_argument("fid", nargs="?", default=None, help="The FID to look up (required for tool and resolve modes)")
    parser.add_argument("--out-dir", default=os.getcwd(), help="Output folder (defaults to the current working directory)")
    parser.add_argument(
        "--mode",
        choices=["tool", "modules", "zbom", "resolve"],
        default="tool",
        help=(
            "tool = Full BOM (IB53, default, requires FID); "
            "modules = split by module (ZOOBOM_CE_FMT, requires SO or FID); "
            "zbom = configuration options (VA03, requires SO or FID); "
            "resolve = look up SO/PO only, no download (requires FID)"
        ),
    )
    parser.add_argument("--so", default=None, help="SO number (required for modules/zbom modes unless FID is given)")
    args = parser.parse_args()

    if args.mode in ("tool", "resolve") and not args.fid:
        log(f"[Error] {args.mode} mode requires a FID.")
        sys.exit(1)

    try:
        fid = args.fid.strip() if args.fid else None
        so = args.so.strip() if args.so else None
        xlsx_path = run(fid, args.out_dir, args.mode, so)
    except Exception as e:
        log(f"[Error] {e}")
        sys.exit(1)

    if xlsx_path:
        log(f"RESULT_PATH:{xlsx_path}")
    sys.exit(0)


if __name__ == "__main__":
    main()
